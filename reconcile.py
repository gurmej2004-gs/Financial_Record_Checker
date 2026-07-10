import csv
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import argparse
import sys


def load_csv_files(bank_file, records_file):
    """Load bank statement and internal records CSV files."""
    try:
        bank_transactions = []
        records_transactions = []
        
        # Read bank statement
        with open(bank_file, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                bank_transactions.append({
                    'transaction_id': row['transaction_id'],
                    'date': datetime.strptime(row['date'], '%Y-%m-%d'),
                    'description': row['description'],
                    'amount': float(row['amount'])
                })
        
        # Read internal records
        with open(records_file, 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                records_transactions.append({
                    'transaction_id': row['transaction_id'],
                    'date': datetime.strptime(row['date'], '%Y-%m-%d'),
                    'description': row['description'],
                    'amount': float(row['amount'])
                })
        
        return bank_transactions, records_transactions
    except Exception as e:
        print(f"Error loading CSV files: {e}")
        sys.exit(1)


def match_transactions(bank_transactions, records_transactions, date_tolerance_days=1):
    """Match transactions between bank statement and records."""
    matched = []
    missing_in_records = []
    missing_in_bank = []
    
    # Track matched transaction IDs
    matched_bank_ids = set()
    matched_records_ids = set()
    
    # Match each bank transaction to records
    for bank_tx in bank_transactions:
        bank_id = bank_tx['transaction_id']
        bank_amount = bank_tx['amount']
        bank_date = bank_tx['date']
        
        # Calculate date range
        date_start = bank_date - timedelta(days=date_tolerance_days)
        date_end = bank_date + timedelta(days=date_tolerance_days)
        
        # Find matching record
        found_match = False
        for records_tx in records_transactions:
            if (records_tx['transaction_id'] not in matched_records_ids and
                records_tx['amount'] == bank_amount and
                date_start <= records_tx['date'] <= date_end):
                
                matched.append({
                    'bank_transaction_id': bank_id,
                    'records_transaction_id': records_tx['transaction_id'],
                    'date': bank_date.strftime('%Y-%m-%d'),
                    'description': bank_tx['description'],
                    'amount': bank_amount,
                    'status': 'Matched'
                })
                matched_bank_ids.add(bank_id)
                matched_records_ids.add(records_tx['transaction_id'])
                found_match = True
                break
        
        if not found_match:
            missing_in_records.append({
                'transaction_id': bank_id,
                'date': bank_date.strftime('%Y-%m-%d'),
                'description': bank_tx['description'],
                'amount': bank_amount,
                'status': 'Missing in Records'
            })
    
    # Find records transactions not matched to bank
    for records_tx in records_transactions:
        if records_tx['transaction_id'] not in matched_records_ids:
            missing_in_bank.append({
                'transaction_id': records_tx['transaction_id'],
                'date': records_tx['date'].strftime('%Y-%m-%d'),
                'description': records_tx['description'],
                'amount': records_tx['amount'],
                'status': 'Missing in Bank Statement'
            })
    
    return matched, missing_in_records, missing_in_bank


def generate_summary(matched, missing_in_records, missing_in_bank):
    """Generate summary statistics."""
    total_transactions = len(matched) + len(missing_in_records) + len(missing_in_bank)
    matched_count = len(matched)
    unmatched_count = len(missing_in_records) + len(missing_in_bank)
    
    total_unmatched_value = 0
    for item in missing_in_records + missing_in_bank:
        total_unmatched_value += item['amount']
    
    return {
        'total_transactions': total_transactions,
        'matched_count': matched_count,
        'unmatched_count': unmatched_count,
        'missing_in_records_count': len(missing_in_records),
        'missing_in_bank_count': len(missing_in_bank),
        'total_unmatched_value': total_unmatched_value
    }


def write_sheet_data(worksheet, data, headers, header_style, row_style=None):
    """Write data to Excel sheet with styling."""
    # Write headers
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = header_style['font']
        cell.fill = header_style['fill']
        cell.alignment = header_style['alignment']
    
    # Write data rows
    for row_data in data:
        row_values = [row_data.get(header, '') for header in headers]
        worksheet.append(row_values)
    
    # Auto-adjust column widths
    for col_num, header in enumerate(headers, 1):
        max_length = len(header)
        for row_data in data:
            value = str(row_data.get(header, ''))
            if len(value) > max_length:
                max_length = len(value)
        worksheet.column_dimensions[chr(64 + col_num)].width = max_length + 2


def create_excel_report(matched, missing_in_records, missing_in_bank, summary, output_file):
    """Create formatted Excel report with multiple sheets."""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    header_style = {
        'font': header_font,
        'fill': header_fill,
        'alignment': header_alignment
    }
    
    unmatched_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Sheet 1: Matched Transactions
    if matched:
        ws_matched = wb.create_sheet('Matched')
        matched_headers = ['bank_transaction_id', 'records_transaction_id', 'date', 'description', 'amount', 'status']
        write_sheet_data(ws_matched, matched, matched_headers, header_style)
    
    # Sheet 2: Discrepancies
    ws_discrepancies = wb.create_sheet('Discrepancies')
    discrepancies = missing_in_records + missing_in_bank
    if discrepancies:
        discrepancies_headers = ['transaction_id', 'date', 'description', 'amount', 'status']
        write_sheet_data(ws_discrepancies, discrepancies, discrepancies_headers, header_style)
        
        # Apply red fill to data rows
        for row in ws_discrepancies.iter_rows(min_row=2):
            for cell in row:
                cell.fill = unmatched_fill
    else:
        ws_discrepancies.append(['No discrepancies found'])
    
    # Sheet 3: Summary
    ws_summary = wb.create_sheet('Summary')
    ws_summary.append(['Metric', 'Value'])
    ws_summary.append(['Total Transactions Processed', summary['total_transactions']])
    ws_summary.append(['Matched Transactions', summary['matched_count']])
    ws_summary.append(['Unmatched Transactions', summary['unmatched_count']])
    ws_summary.append(['Missing in Records', summary['missing_in_records_count']])
    ws_summary.append(['Missing in Bank Statement', summary['missing_in_bank_count']])
    ws_summary.append(['Total Value of Unmatched Transactions', f"${summary['total_unmatched_value']:,.2f}"])
    
    # Format summary header
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 25
    
    # Save workbook
    wb.save(output_file)
    print(f"Excel report saved to: {output_file}")


def main():
    """Main execution function."""
    parser = argparse.ArgumentParser(description='Financial Record Checker')
    parser.add_argument('--bank', default='bank_statement.csv', help='Bank statement CSV file')
    parser.add_argument('--records', default='internal_records.csv', help='Internal records CSV file')
    parser.add_argument('--output', default='bank_reconciliation_report.xlsx', help='Output Excel file')
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("Financial Record Checker")
    print("=" * 60)
    
    # Load data
    print(f"\nLoading bank statement from: {args.bank}")
    print(f"Loading internal records from: {args.records}")
    bank_transactions, records_transactions = load_csv_files(args.bank, args.records)
    
    print(f"Bank statement: {len(bank_transactions)} transactions")
    print(f"Internal records: {len(records_transactions)} transactions")
    
    # Match transactions
    print("\nMatching transactions...")
    matched, missing_in_records, missing_in_bank = match_transactions(bank_transactions, records_transactions)
    
    # Generate summary
    summary = generate_summary(matched, missing_in_records, missing_in_bank)
    
    # Print summary to console
    print("\n" + "=" * 60)
    print("RECONCILIATION SUMMARY")
    print("=" * 60)
    print(f"Total Transactions Processed: {summary['total_transactions']}")
    print(f"Matched Transactions: {summary['matched_count']}")
    print(f"Unmatched Transactions: {summary['unmatched_count']}")
    print(f"  - Missing in Records: {summary['missing_in_records_count']}")
    print(f"  - Missing in Bank Statement: {summary['missing_in_bank_count']}")
    print(f"Total Value of Unmatched Transactions: ${summary['total_unmatched_value']:,.2f}")
    print("=" * 60)
    
    # Create Excel report
    print("\nGenerating Excel report...")
    create_excel_report(matched, missing_in_records, missing_in_bank, summary, args.output)
    
    print("\nReconciliation complete!")


if __name__ == "__main__":
    main()
